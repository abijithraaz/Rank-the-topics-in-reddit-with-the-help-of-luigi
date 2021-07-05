# -*- coding: utf-8 -*-

import praw
import pandas as pd
import openpyxl
import os
import luigi
import time

class MyTask(luigi.Task):
    api_client_id = luigi.Parameter()
    api_client_secret = luigi.Parameter()

    def run(self):
        '''
        write data from reddit
        '''
        # Api connection establishment
        reddit = praw.Reddit(client_id = self.api_client_id, client_secret = self.api_client_secret, user_agent='API_tutorial')

        # Checking the data file is present
        if os.path.isfile('reddit_data.xlsx'):
            os.remove("reddit_data.xlsx")

        '''
        Creation of excel for writing the data
        '''
        wb = openpyxl.Workbook()
        # First sheet for storring posts
        sheet1 = wb.active
        sheet1.title = "POSTS"
        (sheet1.cell(row = 1, column = 1)).value = "PARENTID"
        (sheet1.cell(row = 1, column = 2)).value = "SUBREDDIT"
        (sheet1.cell(row = 1, column = 3)).value = "TITLE"
        (sheet1.cell(row = 1, column = 4)).value = "NOOFCOMMENTS"
        # Second sheet for storing comments
        sheet2 = wb.create_sheet(index = 1 , title = "COMMENTS")
        (sheet2.cell(row = 1, column = 1)).value = "PARENTID"
        (sheet2.cell(row = 1, column = 2)).value = "COMMENTID"
        (sheet2.cell(row = 1, column = 3)).value = "COMMENT"
        (sheet2.cell(row = 1, column = 4)).value = "SCORE"

        # get trendiest posts from all subreddits
        hot_posts = reddit.subreddit('all').hot()
        top_posts = reddit.subreddit('all').top()
        controversial = reddit.subreddit('all').controversial()

        new_list = [hot_posts, top_posts, controversial]    #creating a trending post lists
        i = 2
        k = 2
        for trend in new_list:
            for submission in trend:
                j = 1
                (sheet1.cell(row = i, column = j)).value = submission.id
                j += 1
                (sheet1.cell(row = i, column = j)).value = str(submission.subreddit)
                j += 1
                (sheet1.cell(row = i, column = j)).value = submission.title
                j += 1
                (sheet1.cell(row = i, column = j)).value = int(submission.num_comments)
                i += 1
                if not submission.stickied:
                    submission.comments.replace_more(limit=0)
                    for comment in submission.comments.list():
                        l = 1
                        (sheet2.cell(row = k, column = l)).value = str(comment.parent())
                        l += 1
                        (sheet2.cell(row = k, column = l)).value = comment.id
                        l += 1
                        (sheet2.cell(row = k, column = l)).value = comment.body
                        l += 1
                        (sheet2.cell(row = k, column = l)).value = int(comment.score)
                        k += 1
        wb.save("reddit_data.xlsx")                        #Saving the data

        ''' 
        Get index positions of value in dataframe i.e. dfObj.
        '''
        def getIndexes(dfObj, value):
            
            listOfPos = list()
            # Get bool dataframe with True at positions where the given value exists
            result = dfObj.isin([value])
            # Get list of columns that contains the value
            seriesObj = result.any()
            columnNames = list(seriesObj[seriesObj == True].index)
            # Iterate over list of columns and fetch the rows indexes where value exists
            for col in columnNames:
                rows = list(result[col][result[col] == True].index)
                for row in rows:
                    listOfPos.append(row)
            # Return a list of tuples indicating the positions of value in the dataframe
            return listOfPos
        '''
        Load the data and find the ranks of subbreddit
        '''
        # Reading the excel
        book = openpyxl.load_workbook('reddit_data.xlsx')
        sheet1 = book.active
        sheet2 = book["COMMENTS"]

        # Loading the data to data-frame from excel
        posts = []
        comments = []

        for i, row in enumerate(sheet1.iter_rows()):
            if (i == 0):
                continue
            posts.append([data.value for data in row])
        posts = pd.DataFrame(posts,columns=['ParentID', 'Subreddit', 'Title', 'Noofcomments'])

        for i, row in enumerate(sheet2.iter_rows()):
            if (i == 0):
                continue
            comments.append([data.value for data in row])       
        comments = pd.DataFrame(comments,columns=['ParentID', 'CommentId', 'Comment', 'Score'])

        # Sorting the data frames
        posts = posts.sort_values(by = ["ParentID"],ascending=False)
        posts.reset_index(drop=True,inplace=True)
        comments = comments.sort_values(by = ["ParentID", "Score"], ascending=False)
        comments.reset_index(drop=True,inplace=True)

        # Calculating row lengths of posts and comments
        post_no = posts.shape[0]
        comment_no = comments.shape[0]

        '''
        Calculating the post score
        '''
        post_score = 0.0
        post_score_list = []
        total_score = 0

        for i in range(0, post_no):
            total_score = 0    
            # Get list of index positions i.e. row of all occurrences of ParentID in the dataframe
            listOfPositions = getIndexes(comments, (posts.loc[i][0]))
            for j in range(len(listOfPositions)):
                # Score of the comment in 4th column
                # Calculate post score with help of comments score
                total_score = (total_score + int(comments.loc[listOfPositions[j]][3]))    
            comment_no = posts.loc[i][3]
            post_score = (total_score/comment_no)
            post_score_list.append(post_score)
        posts["Postscore"] = post_score_list
        posts = posts.sort_values(by=["Subreddit", "Postscore"], ascending=False)
        posts.reset_index(drop=True,inplace=True)

        # calculating a lists of subreddits
        subreddit_list = []
        for i in range(0, post_no):
            if posts.loc[i][1] not in subreddit_list:
                subreddit_list.append((posts.loc[i][1]))
        '''
        subreddit score finding and ranking
        '''
        subreddit_len = len(subreddit_list)
        subreddit_score_list = []

        for subreddit in subreddit_list:
            total_post_score = 0
            post_index_list = getIndexes(posts, subreddit)
            subreddit_post_len = len(post_index_list)
            for k in range(subreddit_post_len):
                # postScore of the post in 5th column
                total_post_score = total_post_score + (posts.loc[post_index_list[k]][4])
            
            subreddit_score = (total_post_score/subreddit_post_len)
            subreddit_score_list.append([subreddit, subreddit_score])
        subreddit_score_list = pd.DataFrame(subreddit_score_list,columns=['Subreddit', 'subreddit_score'])
        subreddit_score_list = subreddit_score_list.sort_values(by = ["subreddit_score"],ascending=False)
        subreddit_score_list.reset_index(drop=True,inplace=True)

        '''
        Creating output files and logs
        '''
        # Checking the output directory is exist
        if not os.path.isdir('output_log'):
            os.mkdir("output_log")

        wb1 = openpyxl.Workbook()

        # First sheet for storring subreddits
        sheet1 = wb1.active
        sheet1.title = "SubredditRank"
        (sheet1.cell(row = 1, column = 1)).value = "Subreddit"
        (sheet1.cell(row = 1, column = 2)).value = "Rank"

        # Second sheet for storing posts
        sheet2 = wb1.create_sheet(index = 1 , title = "Posts")
        (sheet2.cell(row = 1, column = 1)).value = "Subreddit"
        (sheet2.cell(row = 1, column = 2)).value = "PostID"
        (sheet2.cell(row = 1, column = 3)).value = "TITLE"

        # Third sheet for storing comments
        sheet3 = wb1.create_sheet(index = 2 , title = "Comments")
        (sheet3.cell(row = 1, column = 1)).value = "PostID"
        (sheet3.cell(row = 1, column = 2)).value = "COMMENT"   
        
        '''
        Creating output log file
        '''

        subreddit_no = subreddit_score_list.shape[0]
        rank = 1
        rank_row = 2
        post_row = 2
        comment_row = 2
        for i in range(subreddit_no):
            # Taking top 50 subreddits
            if (i <= 50):
                rank_column = 1
                (sheet1.cell(row = rank_row, column = rank_column)).value = subreddit_score_list.loc[i][0]
                rank_column += 1
                (sheet1.cell(row = rank_row, column = rank_column)).value = rank
                op_post_index_list = getIndexes(posts, subreddit_score_list.loc[i][0])
                # Taking top 10 posts of subreddits
                for m in range(len(op_post_index_list[:10])):
                    post_column = 1
                    (sheet2.cell(row = post_row, column = post_column)).value = subreddit_score_list.loc[i][0]
                    post_column += 1
                    (sheet2.cell(row = post_row, column = post_column)).value = posts.loc[op_post_index_list[m]][0]
                    post_column += 1
                    (sheet2.cell(row = post_row, column = post_column)).value = posts.loc[op_post_index_list[m]][2]
                    op_comment_index_list = getIndexes(comments, posts.loc[op_post_index_list[m]][0])
                    # Taking top 5 comments of posts
                    for n in range(len(op_comment_index_list[:5])):
                        comment_column = 1
                        (sheet3.cell(row = comment_row, column = comment_column)).value = posts.loc[op_post_index_list[m]][0]
                        comment_column += 1
                        (sheet3.cell(row = comment_row, column = comment_column)).value = comments.loc[op_comment_index_list[n]][2]
                        comment_row += 1        
                    post_row += 1
                rank += 1
                rank_row += 1
            else:
                break
        local_time = time.localtime()
        time_string = time.strftime("%m_%d_%Y_%H_%M_%S", local_time)
        wb1.save("output_log/subreddit_rank_op_{}.xlsx".format(time_string))

if __name__ == "__main__":
    luigi.run()
